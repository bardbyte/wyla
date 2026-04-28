"""Tools the gold_curator agent uses to inspect Excel files + validate SQL.

Design principles:
  - Each tool returns a `dict` with `status: "ok" | "error"` and either
    payload fields or an `error` string. ADK's Gemini reads the function
    signature + docstring to generate the tool schema, so docstrings must
    be precise — they're what the LLM sees when deciding which tool to call.
  - No interpretation here. Tools surface the raw structure (sheet names,
    cell values, sqlglot's parse output) and let the agent reason about it.
  - Markdown tables for previews — Gemini parses them natively, so they're
    the highest-fidelity way to show tabular data without blowing context.
  - Caps and pagination on everything that touches a sheet, so a 50K-row
    workbook can't OOM the model context.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import sqlglot
from openpyxl import load_workbook
from sqlglot import exp

# Cap on how many cells we'll dump into one tool response. ~200 rows × ~30 cols
# of typical gold-query data is well under Gemini's context budget; agents that
# need more should call read_excel_rows in pages.
_MAX_PREVIEW_ROWS = 50
_MAX_FULL_ROWS = 1000
_MAX_CELL_PREVIEW_CHARS = 240
_MAX_DISTINCT_VALUES_TO_SHOW = 12


# --------------------------------------------------------------------------- #
# Excel structure / content                                                   #
# --------------------------------------------------------------------------- #


def list_excel_sheets(file_path: str) -> dict[str, Any]:
    """List every sheet in an Excel workbook with its dimensions.

    Always call this first. The output tells you which sheets exist and how
    big they are, so you can decide which one to preview.

    Args:
        file_path: Absolute or ~-relative path to a .xlsx or .xlsm file.

    Returns:
        dict with:
          status: "ok" | "error"
          file_path: resolved absolute path
          file_size_kb: file size in KB
          sheet_count: int
          sheets: list of {name, rows, cols} — `rows` and `cols` are openpyxl's
              max_row / max_column (may overcount blank trailing area).
          error: str | None
    """
    p = _resolve(file_path)
    if isinstance(p, dict):
        return p

    if p.suffix.lower() not in {".xlsx", ".xlsm"}:
        return _err(
            f"unsupported extension {p.suffix} — convert to .xlsx (openpyxl reads only xlsx/xlsm)"
        )

    try:
        wb = load_workbook(p, read_only=True, data_only=True)
    except Exception as e:
        return _err(f"openpyxl could not open file: {e}")

    sheets = [
        {"name": ws.title, "rows": ws.max_row or 0, "cols": ws.max_column or 0}
        for ws in wb.worksheets
    ]
    wb.close()

    return {
        "status": "ok",
        "file_path": str(p),
        "file_size_kb": round(p.stat().st_size / 1024, 1),
        "sheet_count": len(sheets),
        "sheets": sheets,
        "error": None,
    }


def preview_excel_sheet(
    file_path: str, sheet_name: str, num_rows: int = 10
) -> dict[str, Any]:
    """Read the first N rows of a sheet and return them as a markdown table.

    Use this to see what the data actually looks like — header name alone is
    often misleading. Gemini reads markdown tables natively, so the
    `preview_markdown` field is the most useful thing to inspect.

    Args:
        file_path: Path to the workbook.
        sheet_name: Sheet to preview (must match a name from list_excel_sheets).
        num_rows: How many data rows to include (excluding the header). Capped
            at 50.

    Returns:
        dict with:
          status, sheet, header (list[str]), preview_rows (int returned),
          total_rows (int in sheet incl. header), preview_markdown (str —
          a rendered markdown table), error.
    """
    p = _resolve(file_path)
    if isinstance(p, dict):
        return p
    n = max(1, min(num_rows, _MAX_PREVIEW_ROWS))

    wb, err = _open_sheet(p, sheet_name)
    if err:
        return err
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True, max_row=n + 1))
    total = ws.max_row or 0
    wb.close()

    if not rows:
        return {
            "status": "ok",
            "sheet": sheet_name,
            "header": [],
            "preview_rows": 0,
            "total_rows": 0,
            "preview_markdown": "(empty sheet)",
            "error": None,
        }

    header = [_clean_str(c) or f"(col{i})" for i, c in enumerate(rows[0])]
    body = rows[1:]
    md = _to_markdown_table(header, body)

    return {
        "status": "ok",
        "sheet": sheet_name,
        "header": header,
        "preview_rows": len(body),
        "total_rows": total,
        "preview_markdown": md,
        "error": None,
    }


def read_excel_rows(
    file_path: str,
    sheet_name: str,
    start_row: int,
    end_row: int,
) -> dict[str, Any]:
    """Read a specific row range from a sheet.

    Use this to look deeper than `preview_excel_sheet` allows — e.g., spot-
    check rows 100-110 to see if SQL quality holds up across the whole file.
    Hard-capped at 1000 rows per call.

    Args:
        file_path: Path to the workbook.
        sheet_name: Sheet to read.
        start_row: 1-based first row to include (1 = header).
        end_row: 1-based last row to include (inclusive).

    Returns:
        dict with status, sheet, rows (list[list[str]]), header (always returned
        from row 1, separate from the requested range), markdown (rendered
        table of just the requested rows), error.
    """
    p = _resolve(file_path)
    if isinstance(p, dict):
        return p
    if start_row < 1 or end_row < start_row:
        return _err(f"invalid range start_row={start_row}, end_row={end_row}")
    if end_row - start_row + 1 > _MAX_FULL_ROWS:
        return _err(
            f"range too wide: {end_row - start_row + 1} > {_MAX_FULL_ROWS}; "
            "page in chunks"
        )

    wb, err = _open_sheet(p, sheet_name)
    if err:
        return err
    ws = wb[sheet_name]

    header_row: list[str] = []
    body: list[list[str]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            header_row = [_clean_str(c) or f"(col{j})" for j, c in enumerate(row)]
        if start_row <= i <= end_row:
            body.append([_clean_str(c) for c in row])
        if i > end_row:
            break
    wb.close()

    md = _to_markdown_table(header_row, body) if body else "(no rows in range)"
    return {
        "status": "ok",
        "sheet": sheet_name,
        "header": header_row,
        "row_range": [start_row, end_row],
        "rows_returned": len(body),
        "rows": body,
        "markdown": md,
        "error": None,
    }


def summarize_excel_columns(file_path: str, sheet_name: str) -> dict[str, Any]:
    """Compute per-column statistics for an entire sheet.

    Useful when deciding what each column contains: distinct count, average
    cell length, sample distinct values. Don't rely on header names alone —
    let the data tell you.

    Args:
        file_path: Path to the workbook.
        sheet_name: Sheet to analyze.

    Returns:
        dict with status, sheet, total_rows, columns (list of dicts with
        header, index, non_empty, distinct_count, distinct_sample,
        avg_length, max_length), error.
    """
    p = _resolve(file_path)
    if isinstance(p, dict):
        return p
    wb, err = _open_sheet(p, sheet_name)
    if err:
        return err
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {
            "status": "ok",
            "sheet": sheet_name,
            "total_rows": 0,
            "columns": [],
            "error": None,
        }

    header = [_clean_str(c) or f"(col{i})" for i, c in enumerate(rows[0])]
    n_cols = max(len(header), max((len(r) for r in rows[1:]), default=0))
    header += [f"(col{i})" for i in range(len(header), n_cols)]

    columns: list[dict[str, Any]] = []
    for col_i in range(n_cols):
        values = [_clean_str(r[col_i]) if col_i < len(r) else "" for r in rows[1:]]
        non_empty = [v for v in values if v]
        distinct = list(dict.fromkeys(non_empty))  # preserves order, dedupes
        columns.append(
            {
                "header": header[col_i],
                "index": col_i,
                "non_empty": len(non_empty),
                "distinct_count": len(set(non_empty)),
                "distinct_sample": [
                    _truncate(v, 80) for v in distinct[:_MAX_DISTINCT_VALUES_TO_SHOW]
                ],
                "avg_length": (
                    round(sum(len(v) for v in non_empty) / max(len(non_empty), 1), 1)
                ),
                "max_length": max((len(v) for v in non_empty), default=0),
            }
        )

    return {
        "status": "ok",
        "sheet": sheet_name,
        "total_rows": len(rows) - 1,  # excl. header
        "columns": columns,
        "error": None,
    }


# --------------------------------------------------------------------------- #
# SQL validation + LookML decomposition                                       #
# --------------------------------------------------------------------------- #


def validate_sql(sql: str, dialect: str = "bigquery") -> dict[str, Any]:
    """Parse a SQL query with sqlglot and report syntax + structural metadata.

    Use this to assess whether the gold queries are syntactically clean. If
    `status == "error"`, the query won't be usable downstream and the dataset
    needs fixing.

    Args:
        sql: The SQL text to validate.
        dialect: SQL dialect — default "bigquery". Other valid values:
            "snowflake", "postgres", "mysql", "ansi".

    Returns:
        dict with status, parses (bool), tables (list[str]), join_count,
        agg_count, has_cte, has_window, complexity ("simple"|"medium"|"complex"),
        error (with sqlglot's exact message on failure).
    """
    try:
        tree = sqlglot.parse_one(sql, dialect=dialect)
    except sqlglot.errors.ParseError as e:
        return {
            "status": "error",
            "parses": False,
            "error": f"sqlglot ParseError: {e}",
        }
    except Exception as e:
        return {
            "status": "error",
            "parses": False,
            "error": f"{type(e).__name__}: {e}",
        }

    tables = sorted({t.name for t in tree.find_all(exp.Table) if t.name})
    joins = list(tree.find_all(exp.Join))
    aggs = [
        node
        for node in tree.walk()
        if isinstance(node, exp.AggFunc | exp.Count | exp.Sum | exp.Avg | exp.Min | exp.Max)
    ]
    has_cte = bool(tree.find(exp.With))
    has_window = bool(tree.find(exp.Window))

    if has_cte or has_window or len(joins) >= 2:
        complexity = "complex"
    elif joins or len(aggs) >= 2:
        complexity = "medium"
    else:
        complexity = "simple"

    return {
        "status": "ok",
        "parses": True,
        "tables": tables,
        "join_count": len(joins),
        "agg_count": len(aggs),
        "has_cte": has_cte,
        "has_window": has_window,
        "complexity": complexity,
        "error": None,
    }


def analyze_for_lookml(sql: str, dialect: str = "bigquery") -> dict[str, Any]:
    """Decompose a SQL query into LookML primitives.

    Returns the building blocks an analyst would map into a LookML view:
    base tables, dimensions used, measures (aggregations), filters, joins.
    Use this to assess whether each gold query is representable as a LookML
    semantic-layer query, and what view fields are needed.

    Args:
        sql: SQL text to decompose.
        dialect: SQL dialect — default "bigquery".

    Returns:
        dict with status, tables, primary_table, dimensions (column names
        appearing in SELECT/GROUP BY non-aggregated), measures (list of
        {function, column, expression}), filters (list of {column, operator,
        value}), joins (list of {right_table, side, condition}), lookml_ready
        (bool — heuristic: has tables and at least one measure or dimension),
        error.
    """
    try:
        tree = sqlglot.parse_one(sql, dialect=dialect)
    except sqlglot.errors.ParseError as e:
        return {"status": "error", "error": f"sqlglot ParseError: {e}"}

    # Tables (in FROM order if possible)
    tables_ordered: list[str] = []
    for t in tree.find_all(exp.Table):
        if t.name and t.name not in tables_ordered:
            tables_ordered.append(t.name)

    # Dimensions: non-aggregated SELECT items + GROUP BY columns
    dimensions: list[str] = []
    measures: list[dict[str, Any]] = []
    select = tree.find(exp.Select)
    if select:
        for e in select.expressions or []:
            unaliased = e.unalias() if isinstance(e, exp.Alias) else e
            agg_class = next(
                (
                    cls
                    for cls in (exp.Sum, exp.Count, exp.Avg, exp.Min, exp.Max, exp.Stddev, exp.Variance)
                    if isinstance(unaliased, cls)
                ),
                None,
            )
            if agg_class is not None:
                inner = unaliased.this
                if isinstance(inner, exp.Distinct):
                    inner = (inner.expressions or [inner.this])[0] if (inner.expressions or [inner.this]) else None
                column = inner.name if isinstance(inner, exp.Column) else None
                measures.append(
                    {
                        "function": agg_class.__name__.upper(),
                        "column": column,
                        "expression": unaliased.sql(dialect=dialect),
                    }
                )
            elif isinstance(unaliased, exp.Column) and unaliased.name not in dimensions:
                dimensions.append(unaliased.name)

    for g in tree.find_all(exp.Group):
        for col in g.expressions or []:
            if isinstance(col, exp.Column) and col.name not in dimensions:
                dimensions.append(col.name)

    # Filters
    filters: list[dict[str, str]] = []
    for where in tree.find_all(exp.Where):
        for eq in where.find_all(exp.EQ):
            if isinstance(eq.left, exp.Column):
                filters.append(
                    {
                        "column": eq.left.name,
                        "operator": "=",
                        "value": eq.right.sql(dialect=dialect),
                    }
                )

    # Joins
    joins: list[dict[str, Any]] = []
    for j in tree.find_all(exp.Join):
        right_tbl = j.this.name if isinstance(j.this, exp.Table) else None
        on = j.args.get("on")
        side = (j.side or "").lower()
        kind_raw = (j.kind or "").lower()
        joins.append(
            {
                "right_table": right_tbl,
                "side": side or kind_raw or "inner",
                "condition": on.sql(dialect=dialect) if on is not None else None,
            }
        )

    return {
        "status": "ok",
        "tables": tables_ordered,
        "primary_table": tables_ordered[0] if tables_ordered else None,
        "dimensions": dimensions,
        "measures": measures,
        "filters": filters,
        "joins": joins,
        "lookml_ready": bool(tables_ordered) and (bool(dimensions) or bool(measures)),
        "error": None,
    }


# --------------------------------------------------------------------------- #
# Final extraction                                                            #
# --------------------------------------------------------------------------- #


def extract_gold_queries(
    file_path: str,
    sheet_name: str,
    prompt_column: str,
    sql_column: str,
    difficulty_column: str | None = None,
    id_column: str | None = None,
    output_json_path: str | None = None,
) -> dict[str, Any]:
    """Materialize the gold queries as a structured list once you've decided
    which columns hold what.

    Call this AFTER you've inspected the workbook (list_excel_sheets +
    preview_excel_sheet) and identified the column names. If your inspection
    suggested wrong columns, the response includes `available_columns` so
    you can self-correct.

    Args:
        file_path: Path to the workbook.
        sheet_name: Sheet containing the queries.
        prompt_column: Header name of the natural-language prompt column.
        sql_column: Header name of the expected-SQL column.
        difficulty_column: Optional header name for difficulty/tier.
        id_column: Optional header name for an existing ID column. If unset,
            we generate q_0001, q_0002, ...
        output_json_path: Optional path to save the extracted records as JSON.
            Saved as utf-8 with indent=2.

    Returns:
        dict with status, extracted (count), prompt_col, sql_col,
        difficulty_col, id_col, sample (first 3 records),
        rows_skipped_empty, saved_to, available_columns (on error), error.
    """
    p = _resolve(file_path)
    if isinstance(p, dict):
        return p
    wb, err = _open_sheet(p, sheet_name)
    if err:
        return err
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return _err(f"sheet '{sheet_name}' is empty")

    header = [_clean_str(c) for c in rows[0]]

    def idx(name: str | None) -> int | None:
        if not name:
            return None
        try:
            return header.index(name)
        except ValueError:
            return None

    pi, si = idx(prompt_column), idx(sql_column)
    di, ii = idx(difficulty_column), idx(id_column)
    if pi is None or si is None:
        return {
            "status": "error",
            "error": (
                f"missing columns: prompt_column={prompt_column!r} (idx={pi}), "
                f"sql_column={sql_column!r} (idx={si})"
            ),
            "available_columns": header,
        }

    queries: list[dict[str, Any]] = []
    skipped = 0
    for row_n, row in enumerate(rows[1:], start=2):
        prompt = _clean_str(row[pi]) if pi < len(row) else ""
        sql = _clean_str(row[si]) if si < len(row) else ""
        if not prompt and not sql:
            skipped += 1
            continue
        diff = _clean_str(row[di]) if di is not None and di < len(row) else None
        natural_id = _clean_str(row[ii]) if ii is not None and ii < len(row) else ""
        gen_id = natural_id or f"q_{len(queries) + 1:04d}"
        queries.append(
            {
                "id": gen_id,
                "prompt": prompt,
                "sql": sql,
                "difficulty": diff or None,
                "source_row": row_n,
            }
        )

    saved_to: str | None = None
    if output_json_path:
        out_path = Path(output_json_path).expanduser().resolve()
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(queries, indent=2), encoding="utf-8")
        saved_to = str(out_path)

    return {
        "status": "ok",
        "extracted": len(queries),
        "prompt_col": prompt_column,
        "sql_col": sql_column,
        "difficulty_col": difficulty_column,
        "id_col": id_column,
        "sample": queries[:3],
        "rows_skipped_empty": skipped,
        "saved_to": saved_to,
        "error": None,
    }


# --------------------------------------------------------------------------- #
# Helpers                                                                     #
# --------------------------------------------------------------------------- #


def _resolve(file_path: str) -> Path | dict[str, Any]:
    """Return resolved path or an error dict the tool can return directly."""
    p = Path(file_path).expanduser().resolve()
    if not p.exists():
        return _err(f"file not found: {p}")
    return p


def _open_sheet(
    p: Path, sheet_name: str
) -> tuple[Any, dict[str, Any] | None]:
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
    except Exception as e:
        return None, _err(f"openpyxl could not open file: {e}")
    if sheet_name not in [ws.title for ws in wb.worksheets]:
        names = [ws.title for ws in wb.worksheets]
        wb.close()
        return None, {
            "status": "error",
            "error": f"sheet {sheet_name!r} not found",
            "available_sheets": names,
        }
    return wb, None


def _clean_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def _truncate(s: str, n: int) -> str:
    return s if len(s) <= n else s[: n - 1] + "…"


def _to_markdown_table(header: list[str], body_rows: list[Any]) -> str:
    """Render header + rows as a GitHub-flavored markdown table.

    Truncates cells that are too long, escapes pipe chars so the table doesn't
    break, and replaces newlines with spaces so each cell is one line.
    """
    if not header:
        return "(empty)"

    def cell(v: Any) -> str:
        s = _clean_str(v).replace("\n", " ").replace("|", "\\|")
        return _truncate(s, _MAX_CELL_PREVIEW_CHARS)

    lines = ["| " + " | ".join(cell(h) for h in header) + " |"]
    lines.append("| " + " | ".join(["---"] * len(header)) + " |")
    for row in body_rows:
        cells = [cell(row[i] if i < len(row) else "") for i in range(len(header))]
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines)


def _err(msg: str) -> dict[str, Any]:
    return {"status": "error", "error": msg}
