#!/usr/bin/env python3
"""Inspect any Excel sheet, identify which columns hold NL prompts vs SQL vs
difficulty, and extract the gold-query corpus in a form an LLM can consume.

Designed to be column-name agnostic: it doesn't care if your headers say
"user_prompt" or "Question" or "Natural Language Query" or are in Russian —
it scores columns by *content*, not header name. Header names are a tiebreaker
hint, not a requirement.

Standalone usage:

    pip install openpyxl
    python scripts/inspect_excel.py data/gold_queries.xlsx                  # inventory
    python scripts/inspect_excel.py data/gold_queries.xlsx --extract        # + extract
    python scripts/inspect_excel.py data/gold_queries.xlsx --extract --json # machine-readable
    python scripts/inspect_excel.py data/foo.xlsx --extract \\
        --prompt-col 'Question' --sql-col 'Query'                            # override

Two tool-ready functions inside, both return the {status, ..., error} dict
shape we'll lift into lumi/tools/excel_tools.py:

    inspect_excel(path)                         → full inventory + heuristics
    extract_gold_queries(path, ...)             → just the (id, prompt, sql, ...) rows
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    print(
        "ERROR: openpyxl is required. Install with:  pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(2)

DEFAULT_FIXTURE_PATH = Path("tests/fixtures/extracted_gold_queries.json")

# Pattern signatures used to score what each column probably contains.
_SQL_STRONG = re.compile(r"\bSELECT\b.*\bFROM\b", re.IGNORECASE | re.DOTALL)
_SQL_KEYWORDS = re.compile(
    r"\b(SELECT|FROM|WHERE|JOIN|GROUP\s+BY|ORDER\s+BY|HAVING|UNION|WITH|CASE\s+WHEN|INNER|LEFT|RIGHT|FULL)\b",
    re.IGNORECASE,
)
_NL_LEAD_WORDS = {
    "how", "what", "when", "where", "why", "which", "who",
    "show", "count", "list", "find", "get", "give", "tell",
    "average", "total", "sum", "compare", "rank", "calculate",
}
_DIFFICULTY_VOCAB = {
    "easy", "medium", "hard", "extra", "extra hard", "extrahard",
    "low", "moderate", "high", "very hard", "veryhard",
    "simple", "complex", "trivial", "advanced",
    "1", "2", "3", "4", "5", "l1", "l2", "l3",
}
_HEADER_HINT_PROMPT = re.compile(
    r"(question|prompt|query|nl|natural|english|user[_\s]?question|description|ask)",
    re.IGNORECASE,
)
_HEADER_HINT_SQL = re.compile(
    r"(sql|query|expected[_\s]?(sql|query)|answer|gold[_\s]?(sql|query))",
    re.IGNORECASE,
)
_HEADER_HINT_DIFFICULTY = re.compile(
    r"(difficult|complex|level|tier|grade|hardness)", re.IGNORECASE
)
_HEADER_HINT_ID = re.compile(r"(^|[^a-z])(id|key|num|no|index|#|ix)([^a-z]|$)", re.IGNORECASE)


# --------------------------------------------------------------------------- #
# Tool-shaped core                                                            #
# --------------------------------------------------------------------------- #


def inspect_excel(path: str | Path) -> dict[str, Any]:
    """Inventory every sheet of an .xlsx and score each column by likely role.

    Returns:
        {
          "status": "success" | "error",
          "path": str,
          "sheets": [
            {
              "name": str,
              "dimensions": {"rows": int, "cols": int},
              "header_row_index": int (1-based) | None,
              "headers": [str, ...],
              "columns": [
                {
                  "header": str,
                  "index": int (0-based),
                  "non_empty_rows": int,
                  "distinct_count": int,
                  "avg_length": float,
                  "max_length": int,
                  "sample_values": [str, ...],   # up to 3
                  "scores": {sql, nl_prompt, difficulty, id},
                  "detected_role": str | None,
                }
              ],
              "row_count": int,
              "looks_like_gold_queries_score": float,
            }, ...
          ],
          "recommended_sheet": str | None,
          "recommended_columns": {prompt: str|None, sql: str|None, difficulty: str|None, id: str|None},
          "error": str | None,
        }
    """
    p = Path(path)
    if not p.exists():
        return _err(str(p), f"file not found: {p}")
    if p.suffix.lower() not in {".xlsx", ".xlsm"}:
        return _err(
            str(p),
            f"unsupported extension {p.suffix} — convert to .xlsx (openpyxl only reads xlsx/xlsm)",
        )

    try:
        wb: Workbook = load_workbook(p, read_only=True, data_only=True)
    except Exception as e:
        return _err(str(p), f"openpyxl could not open file: {e}")

    sheets_out: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        sheets_out.append(_inspect_sheet(ws))

    # Pick the sheet most likely to be gold queries: top SQL+NL combined score.
    recommended_sheet: str | None = None
    if sheets_out:
        best = max(sheets_out, key=lambda s: s["looks_like_gold_queries_score"])
        if best["looks_like_gold_queries_score"] > 0.2:
            recommended_sheet = best["name"]

    rec_cols: dict[str, str | None] = {
        "prompt": None,
        "sql": None,
        "difficulty": None,
        "id": None,
    }
    # Map column-role names → recommended-column slot names. Sources can use
    # 'nl_prompt' (verbose) but the slot is 'prompt' (terse).
    role_to_slot = {"nl_prompt": "prompt", "sql": "sql", "difficulty": "difficulty", "id": "id"}
    if recommended_sheet:
        rec = next(s for s in sheets_out if s["name"] == recommended_sheet)
        for col in rec["columns"]:
            role = col.get("detected_role")
            slot = role_to_slot.get(role) if role else None
            if slot and rec_cols[slot] is None:
                rec_cols[slot] = col["header"]

    wb.close()
    return {
        "status": "success",
        "path": str(p),
        "sheets": sheets_out,
        "recommended_sheet": recommended_sheet,
        "recommended_columns": rec_cols,
        "error": None,
    }


def extract_gold_queries(
    path: str | Path,
    sheet: str | None = None,
    prompt_col: str | None = None,
    sql_col: str | None = None,
    difficulty_col: str | None = None,
    id_col: str | None = None,
    id_prefix: str = "q_",
) -> dict[str, Any]:
    """Extract (id, prompt, sql, difficulty?) rows from any Excel sheet.

    If column names are not provided, uses inspect_excel() heuristics to pick.
    """
    inv = inspect_excel(path)
    if inv["status"] != "success":
        return {**_err(str(path), inv["error"]), "queries": []}

    target_sheet = sheet or inv.get("recommended_sheet")
    if not target_sheet:
        return {
            **_err(str(path), "no sheet detected as gold-query-shaped; pass --sheet to force"),
            "queries": [],
        }

    sheet_info = next((s for s in inv["sheets"] if s["name"] == target_sheet), None)
    if sheet_info is None:
        return {**_err(str(path), f"sheet '{target_sheet}' not found"), "queries": []}

    # Resolve column names — explicit args win, else the heuristic recs.
    rec = inv["recommended_columns"] if target_sheet == inv.get("recommended_sheet") else {}
    prompt = prompt_col or rec.get("prompt")
    sql = sql_col or rec.get("sql")
    difficulty = difficulty_col or rec.get("difficulty")
    qid = id_col or rec.get("id")

    if not prompt or not sql:
        return {
            **_err(
                str(path),
                "could not identify prompt + SQL columns. Pass --prompt-col and --sql-col explicitly.",
            ),
            "queries": [],
            "sheet": target_sheet,
            "available_headers": sheet_info["headers"],
        }

    # Now actually read the sheet (we already did once but compactly; reread for full rows).
    p = Path(path)
    wb: Workbook = load_workbook(p, read_only=True, data_only=True)
    ws: Worksheet = wb[target_sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_row_idx = (sheet_info["header_row_index"] or 1) - 1
    headers = [_clean_str(c) for c in rows[header_row_idx]] if rows else []

    def _idx(name: str | None) -> int | None:
        if name is None:
            return None
        try:
            return headers.index(name)
        except ValueError:
            return None

    pi, si, di, ii = _idx(prompt), _idx(sql), _idx(difficulty), _idx(qid)
    if pi is None or si is None:
        return {
            **_err(
                str(path),
                f"resolved prompt='{prompt}' (idx={pi}), sql='{sql}' (idx={si}) — column missing in sheet",
            ),
            "queries": [],
            "sheet": target_sheet,
        }

    queries: list[dict[str, Any]] = []
    skipped_empty = 0
    for row_num, row in enumerate(rows[header_row_idx + 1 :], start=header_row_idx + 2):
        prompt_val = _clean_str(row[pi]) if pi < len(row) else ""
        sql_val = _clean_str(row[si]) if si < len(row) else ""
        if not prompt_val and not sql_val:
            skipped_empty += 1
            continue
        diff_val = _clean_str(row[di]) if di is not None and di < len(row) else None
        natural_id = _clean_str(row[ii]) if ii is not None and ii < len(row) else ""
        gen_id = natural_id or f"{id_prefix}{len(queries) + 1:04d}"

        queries.append(
            {
                "id": gen_id,
                "prompt": prompt_val,
                "sql": sql_val,
                "difficulty": diff_val or None,
                "source_sheet": target_sheet,
                "source_row": row_num,
            }
        )

    return {
        "status": "success",
        "path": str(path),
        "sheet": target_sheet,
        "prompt_col": prompt,
        "sql_col": sql,
        "difficulty_col": difficulty,
        "id_col": qid,
        "queries": queries,
        "rows_skipped_empty": skipped_empty,
        "total_rows_after_header": len(rows) - (header_row_idx + 1),
        "error": None,
    }


# --------------------------------------------------------------------------- #
# Sheet inspection internals                                                  #
# --------------------------------------------------------------------------- #


def _inspect_sheet(ws: Worksheet) -> dict[str, Any]:
    rows = list(ws.iter_rows(values_only=True))
    name = ws.title
    if not rows:
        return {
            "name": name,
            "dimensions": {"rows": 0, "cols": 0},
            "header_row_index": None,
            "headers": [],
            "columns": [],
            "row_count": 0,
            "looks_like_gold_queries_score": 0.0,
        }

    header_idx_0based = _detect_header_row(rows)
    if header_idx_0based is None:
        # Treat row 0 as header but flag low confidence by leaving column scoring
        # to operate on the whole column.
        header_idx_0based = 0

    headers = [_clean_str(c) for c in rows[header_idx_0based]]
    data_rows = rows[header_idx_0based + 1 :]
    n_cols = max(len(headers), max((len(r) for r in data_rows), default=0))
    headers = headers + [""] * (n_cols - len(headers))

    columns: list[dict[str, Any]] = []
    for col_i in range(n_cols):
        values = [
            _clean_str(r[col_i]) if col_i < len(r) else ""
            for r in data_rows
        ]
        non_empty = [v for v in values if v]
        scores = _score_column(values, headers[col_i])
        detected = _pick_role(scores)
        columns.append(
            {
                "header": headers[col_i] or f"(col{col_i})",
                "index": col_i,
                "non_empty_rows": len(non_empty),
                "distinct_count": len({v for v in non_empty}),
                "avg_length": (
                    round(sum(len(v) for v in non_empty) / max(len(non_empty), 1), 1)
                ),
                "max_length": max((len(v) for v in non_empty), default=0),
                "sample_values": [
                    _truncate(v, 100) for v in non_empty[:3]
                ],
                "scores": {k: round(v, 3) for k, v in scores.items()},
                "detected_role": detected,
            }
        )

    # Sheet-level "is this gold queries?" score = best SQL × best NL.
    best_sql = max((c["scores"]["sql"] for c in columns), default=0)
    best_nl = max((c["scores"]["nl_prompt"] for c in columns), default=0)
    sheet_score = round((best_sql + best_nl) / 2, 3)

    return {
        "name": name,
        "dimensions": {"rows": len(rows), "cols": n_cols},
        "header_row_index": header_idx_0based + 1,  # 1-based for human
        "headers": headers,
        "columns": columns,
        "row_count": len(data_rows),
        "looks_like_gold_queries_score": sheet_score,
    }


def _detect_header_row(rows: list[tuple[Any, ...]]) -> int | None:
    """Find the row that's most likely the header.

    Heuristic: the first row whose cells are mostly non-empty short strings,
    AND the row immediately below has substantially different characteristics
    (longer, more varied content). Falls back to row 0.

    Skips leading blank rows (notes, banners) up to 5 deep.
    """
    for i, row in enumerate(rows[:6]):
        cells = [_clean_str(c) for c in row]
        non_empty = [c for c in cells if c]
        if not non_empty:
            continue
        # Headers: ≥ 2 short string cells, all under 60 chars, no SQL keywords.
        all_strings_short = all(0 < len(c) < 60 for c in non_empty)
        no_sql = all(not _SQL_KEYWORDS.search(c) for c in non_empty)
        if len(non_empty) >= 2 and all_strings_short and no_sql:
            # Confirm the next row looks like data (longer or more varied).
            if i + 1 < len(rows):
                next_cells = [_clean_str(c) for c in rows[i + 1]]
                next_non_empty = [c for c in next_cells if c]
                if next_non_empty and any(len(c) > 40 for c in next_non_empty):
                    return i
                # Also accept if the next row has a non-string-looking cell.
                return i
            return i
    return None


def _score_column(values: list[str], header: str) -> dict[str, float]:
    """Score how likely this column is each role: sql, nl_prompt, difficulty, id."""
    non_empty = [v for v in values if v]
    if not non_empty:
        return {"sql": 0.0, "nl_prompt": 0.0, "difficulty": 0.0, "id": 0.0}
    n = len(non_empty)

    # --- SQL ---
    strong_sql = sum(1 for v in non_empty if _SQL_STRONG.search(v))
    weak_sql = sum(1 for v in non_empty if _SQL_KEYWORDS.search(v))
    sql_score = min(1.0, (strong_sql + 0.4 * (weak_sql - strong_sql)) / n)

    # --- NL prompt ---
    # Only bail on cells that look like *actual* SQL (SELECT...FROM). Plain
    # English sentences contain "from" / "where" / "join" all the time —
    # penalizing those was hiding real NL columns.
    nl_total = 0.0
    for v in non_empty:
        v_lower = v.lower().strip()
        if not v_lower or _SQL_STRONG.search(v):
            continue
        s = 0.0
        first = v_lower.split()[0] if v_lower.split() else ""
        if first in _NL_LEAD_WORDS:
            s += 0.45
        if v_lower.endswith("?"):
            s += 0.35
        if 12 <= len(v) <= 350:
            s += 0.15
        if " " in v_lower:  # multi-word
            s += 0.05
        # Bonus: NL prompts have SQL-keyword density ≪ real SQL.
        kw_hits = len(_SQL_KEYWORDS.findall(v))
        if kw_hits == 0:
            s += 0.10
        nl_total += min(1.0, s)
    nl_score = nl_total / n

    # --- Difficulty ---
    distinct_lower = {v.lower().strip() for v in non_empty}
    if 0 < len(distinct_lower) <= 8:
        in_vocab = sum(1 for d in distinct_lower if d in _DIFFICULTY_VOCAB)
        difficulty_score = in_vocab / max(len(distinct_lower), 1)
    else:
        difficulty_score = 0.0

    # --- ID ---
    int_like = sum(1 for v in non_empty if v.strip().isdigit())
    uuid_like = sum(
        1
        for v in non_empty
        if re.fullmatch(r"[a-fA-F0-9]{8}-[a-fA-F0-9]{4}-[a-fA-F0-9]{4}-[a-fA-F0-9]{4}-[a-fA-F0-9]{12}", v.strip())
    )
    short_id_like = sum(
        1
        for v in non_empty
        if re.fullmatch(r"[a-zA-Z][a-zA-Z0-9_]{0,30}", v.strip()) and len(v) < 25
    )
    id_score = max(int_like, uuid_like, short_id_like * 0.5) / n

    # --- Header-name boost (small, just a tiebreaker) ---
    if header:
        if _HEADER_HINT_SQL.search(header):
            sql_score = min(1.0, sql_score + 0.1)
        if _HEADER_HINT_PROMPT.search(header):
            nl_score = min(1.0, nl_score + 0.1)
        if _HEADER_HINT_DIFFICULTY.search(header):
            difficulty_score = min(1.0, difficulty_score + 0.2)
        if _HEADER_HINT_ID.search(header):
            id_score = min(1.0, id_score + 0.15)

    return {
        "sql": sql_score,
        "nl_prompt": nl_score,
        "difficulty": difficulty_score,
        "id": id_score,
    }


def _pick_role(scores: dict[str, float]) -> str | None:
    """Pick the role this column best fits, if any score is meaningfully high."""
    threshold = 0.35
    role = max(scores.items(), key=lambda kv: kv[1])
    name, val = role
    return name if val >= threshold else None


def _clean_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def _truncate(s: str, n: int) -> str:
    return s if len(s) <= n else s[: n - 1] + "…"


def _err(path: str, error: str) -> dict[str, Any]:
    return {
        "status": "error",
        "path": path,
        "sheets": [],
        "recommended_sheet": None,
        "recommended_columns": {"prompt": None, "sql": None, "difficulty": None, "id": None},
        "error": error,
    }


# --------------------------------------------------------------------------- #
# Pretty-printing                                                             #
# --------------------------------------------------------------------------- #


def _format_inventory(inv: dict[str, Any]) -> str:
    if inv["status"] != "success":
        return f"FAILED: {inv['error']}"

    lines: list[str] = []
    lines.append(f"File: {inv['path']}")
    lines.append(f"Sheets: {len(inv['sheets'])}")
    lines.append("")

    for s in inv["sheets"]:
        marker = " ← recommended" if s["name"] == inv["recommended_sheet"] else ""
        lines.append(
            f"[Sheet: {s['name']}]{marker}"
            f"  rows={s['dimensions']['rows']}, cols={s['dimensions']['cols']}, "
            f"data_rows={s['row_count']}, "
            f"gold_query_score={s['looks_like_gold_queries_score']}"
        )
        lines.append(
            f"  Header row: {s['header_row_index'] or '(unknown — using row 1)'}"
        )
        lines.append("  Columns:")
        for col in s["columns"]:
            role = col["detected_role"] or "—"
            lines.append(
                f"    [{col['index']}] {col['header']:<30}  "
                f"role={role:<11}  "
                f"scores=sql:{col['scores']['sql']:.2f} "
                f"nl:{col['scores']['nl_prompt']:.2f} "
                f"diff:{col['scores']['difficulty']:.2f} "
                f"id:{col['scores']['id']:.2f}  "
                f"non_empty={col['non_empty_rows']}, "
                f"distinct={col['distinct_count']}, "
                f"avg_len={col['avg_length']}"
            )
            for sample in col["sample_values"]:
                lines.append(f"        sample: {_truncate(sample, 90)}")
        lines.append("")

    lines.append("Recommended extraction:")
    lines.append(f"  sheet:      {inv['recommended_sheet'] or '(none — pass --sheet)'}")
    rc = inv["recommended_columns"]
    lines.append(f"  prompt_col: {rc['prompt'] or '(unknown — pass --prompt-col)'}")
    lines.append(f"  sql_col:    {rc['sql'] or '(unknown — pass --sql-col)'}")
    lines.append(f"  diff_col:   {rc['difficulty'] or '(none)'}")
    lines.append(f"  id_col:     {rc['id'] or '(none — auto-generate)'}")
    return "\n".join(lines)


def _format_extraction(ext: dict[str, Any]) -> str:
    if ext["status"] != "success":
        out = f"FAILED: {ext['error']}"
        if "available_headers" in ext:
            out += f"\nAvailable headers in '{ext.get('sheet', '?')}': {ext['available_headers']}"
        return out
    qs = ext["queries"]
    lines = [
        f"Extracted {len(qs)} queries from sheet '{ext['sheet']}'",
        f"  prompt_col: {ext['prompt_col']}",
        f"  sql_col:    {ext['sql_col']}",
        f"  diff_col:   {ext.get('difficulty_col') or '(none)'}",
        f"  rows skipped (blank): {ext['rows_skipped_empty']}",
        "",
        "First 3 sample queries:",
    ]
    for q in qs[:3]:
        lines.append(f"  [{q['id']}] (row {q['source_row']}, difficulty={q.get('difficulty')})")
        lines.append(f"    prompt: {_truncate(q['prompt'], 140)}")
        lines.append(f"    sql:    {_truncate(q['sql'], 140)}")
    return "\n".join(lines)


# --------------------------------------------------------------------------- #
# CLI                                                                         #
# --------------------------------------------------------------------------- #


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="inspect_excel",
        description="Inspect any Excel and (optionally) extract gold queries.",
    )
    parser.add_argument("path", help="Path to .xlsx file")
    parser.add_argument(
        "--extract",
        action="store_true",
        help="Run the extraction (otherwise just inventory).",
    )
    parser.add_argument("--sheet", help="Force a specific sheet.")
    parser.add_argument("--prompt-col", help="Override detected prompt column name.")
    parser.add_argument("--sql-col", help="Override detected SQL column name.")
    parser.add_argument("--difficulty-col", help="Override detected difficulty column.")
    parser.add_argument("--id-col", help="Use this column for query IDs (else auto-generate).")
    parser.add_argument(
        "--save",
        default=str(DEFAULT_FIXTURE_PATH),
        help=f"Where to save extracted queries (with --extract). Default: {DEFAULT_FIXTURE_PATH}",
    )
    parser.add_argument("--no-save", action="store_true")
    parser.add_argument("--json", action="store_true", help="Print JSON.")
    args = parser.parse_args(argv)

    if not args.extract:
        result = inspect_excel(args.path)
        if args.json:
            print(json.dumps(result, indent=2, default=str))
        else:
            print(_format_inventory(result))
        return 0 if result["status"] == "success" else 1

    result = extract_gold_queries(
        args.path,
        sheet=args.sheet,
        prompt_col=args.prompt_col,
        sql_col=args.sql_col,
        difficulty_col=args.difficulty_col,
        id_col=args.id_col,
    )

    if args.json:
        print(json.dumps(result, indent=2, default=str))
    else:
        print(_format_extraction(result))

    if not args.no_save and result["status"] == "success":
        target = Path(args.save)
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(json.dumps(result, indent=2, default=str), encoding="utf-8")
        print()
        print(f"Saved {len(result['queries'])} queries → {target}")

    return 0 if result["status"] == "success" else 1


if __name__ == "__main__":
    raise SystemExit(main())
