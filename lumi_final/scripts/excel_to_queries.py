#!/usr/bin/env python3
"""Read SQL queries out of an Excel column and write them as the LUMI inputs.

Two outputs from one input:
  1. data/gold_queries/Q001.sql ... Q129.sql  (one .sql per row — pipeline input)
  2. data/gold_queries/queries.json           (consolidated, JSON-shaped, with
                                              prompt/SQL/difficulty/derived metadata)

Optional Gemini cleanup pass on each row:
  - normalizes smart-quotes / line-endings
  - strips inline comments and surrounding markdown fences
  - flags rows the model thinks are non-SQL (notes, blanks, broken syntax)

Usage (no LLM):
    python scripts/excel_to_queries.py path/to/queries.xlsx \\
        --sheet "Gold Queries" \\
        --sql-col "Expected SQL" \\
        --prompt-col "Question" \\
        --difficulty-col "Difficulty"

Usage (with Gemini cleanup):
    source agent_test/setup_vertex_env.sh ~/Downloads/key.json
    python scripts/excel_to_queries.py path/to/queries.xlsx \\
        --sql-col "Expected SQL" \\
        --gemini-clean

Auto-detection: if you only know the SQL column, the script will scan the
sheet header and ask which one to use. With --auto you get a heuristic
(longest avg-length non-empty column wins for SQL).

Output schema (queries.json), one record per row:
    {
      "id": "Q001",
      "prompt": "What is the total billed business for cornerstone in Q1 2025?",
      "sql": "SELECT SUM(billed_business) FROM ...",
      "difficulty": "easy",
      "source_row": 2,
      "source_sheet": "Gold Queries",
      "source_file": "queries.xlsx",
      "tables": ["cornerstone_metrics"],     # via sqlglot
      "complexity": "simple",                 # simple|medium|complex
      "parse_error": null,                    # str or null
      "cleaned_by_gemini": false              # true if --gemini-clean modified the SQL
    }
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))


# ─── Excel reading (openpyxl) ────────────────────────────────


def load_workbook_or_fail(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        sys.stderr.write(
            f"ERROR: openpyxl not installed. pip install openpyxl. ({e})\n"
        )
        raise SystemExit(2)
    if not path.exists():
        sys.stderr.write(f"ERROR: file not found: {path}\n")
        raise SystemExit(2)
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        sys.stderr.write(
            f"ERROR: unsupported extension {path.suffix}. Convert to .xlsx.\n"
        )
        raise SystemExit(2)
    return load_workbook(path, read_only=True, data_only=True)


def list_sheets(wb) -> list[tuple[str, int, int]]:
    """[(name, max_row, max_col)] for every sheet."""
    return [(ws.title, ws.max_row or 0, ws.max_column or 0) for ws in wb.worksheets]


def read_header(ws) -> list[str]:
    rows = ws.iter_rows(values_only=True, max_row=1)
    first = next(rows, None)
    if first is None:
        return []
    return [_clean(c) or f"(col{i})" for i, c in enumerate(first)]


def _clean(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def col_index(header: list[str], name: str | None) -> int | None:
    if not name:
        return None
    try:
        return header.index(name)
    except ValueError:
        # Case-insensitive fallback
        lname = name.lower()
        for i, h in enumerate(header):
            if h.lower() == lname:
                return i
        return None


def auto_pick_sql_column(ws, header: list[str]) -> int | None:
    """Heuristic: scan first ~25 rows, score each column by 'looks SQL-ish'.

    Score = fraction of cells matching SELECT...FROM. Returns the winning
    column index, or None if nothing scored above threshold.
    """
    SQL_RE = re.compile(r"\bselect\b.*\bfrom\b", re.IGNORECASE | re.DOTALL)
    samples = list(ws.iter_rows(values_only=True, min_row=2, max_row=26))
    if not samples or not header:
        return None
    scores = [0] * len(header)
    counts = [0] * len(header)
    for row in samples:
        for i in range(min(len(row), len(header))):
            v = _clean(row[i])
            if not v:
                continue
            counts[i] += 1
            if SQL_RE.search(v):
                scores[i] += 1
    ratios = [
        (scores[i] / counts[i]) if counts[i] > 0 else 0 for i in range(len(header))
    ]
    best_i = max(range(len(ratios)), key=lambda i: ratios[i]) if ratios else None
    if best_i is None or ratios[best_i] < 0.3:
        return None
    return best_i


# ─── SQL cleaning ────────────────────────────────────────────


_SMART_QUOTES = {
    "“": '"', "”": '"', "‘": "'", "’": "'",
    " ": " ",   # non-breaking space
}

_FENCE_RE = re.compile(r"^\s*```(?:sql|bigquery)?\s*\n(.*?)\n```\s*$", re.DOTALL | re.IGNORECASE)
_LINE_COMMENT_RE = re.compile(r"--[^\n]*$", re.MULTILINE)
_BLOCK_COMMENT_RE = re.compile(r"/\*.*?\*/", re.DOTALL)


def basic_clean(sql: str) -> str:
    """Pure-Python cleanup. No LLM. Always run before Gemini cleanup."""
    s = sql or ""
    # Normalize smart quotes / nbsp
    for bad, good in _SMART_QUOTES.items():
        s = s.replace(bad, good)
    # Strip surrounding markdown fences
    m = _FENCE_RE.match(s.strip())
    if m:
        s = m.group(1)
    # Trim
    return s.strip()


def gemini_clean(sql: str, model_client) -> tuple[str, bool]:
    """Optional LLM normalization. Returns (cleaned_sql, was_modified).

    The model is told to: strip comments, normalize whitespace, ensure the
    output is a SINGLE valid BigQuery SELECT statement, and return the
    cleaned SQL with no commentary. If the row isn't really SQL, return an
    empty string.
    """
    instruction = (
        "You are a SQL janitor. Given the input, return a CLEANED, single "
        "BigQuery SELECT/WITH statement with:\n"
        "  - no markdown fences\n"
        "  - no inline -- or /* */ comments\n"
        "  - normalized whitespace\n"
        "  - smart quotes converted to ASCII\n"
        "If the input is not actually a SQL query (e.g. notes, an empty "
        "string, a description), return EXACTLY the literal token: NOT_SQL\n"
        "Return ONLY the cleaned SQL or NOT_SQL — no preamble, no fences.\n\n"
        "INPUT:\n"
        f"{sql}\n"
    )
    try:
        from google.genai import types

        resp = model_client.models.generate_content(
            model="gemini-3.1-pro-preview",
            contents=instruction,
            config=types.GenerateContentConfig(temperature=0.0),
        )
        text = (resp.text or "").strip()
    except Exception as e:
        sys.stderr.write(f"WARN: Gemini call failed — keeping original. {e}\n")
        return sql, False

    if not text or text == "NOT_SQL":
        return "", True
    # Strip any accidental fence the model emitted anyway
    text = basic_clean(text)
    return text, text != sql


def make_gemini_client():
    """Vertex AI direct (per parent CLAUDE.md). Reuses GOOGLE_* env vars."""
    try:
        import truststore
        truststore.inject_into_ssl()
    except ImportError:
        pass
    from google import genai

    return genai.Client(vertexai=True)


# ─── sqlglot enrichment ──────────────────────────────────────


def derive_metadata(sql: str) -> dict[str, Any]:
    """Use lumi.sql_to_context to extract tables + complexity.

    Same parser the pipeline uses — single source of truth.
    """
    from lumi.sql_to_context import parse_sqls

    fps = parse_sqls([sql])
    fp = fps[0]
    if fp.parse_error:
        return {
            "tables": [],
            "complexity": "unknown",
            "parse_error": fp.parse_error,
        }
    tables = sorted(set(fp.tables) | {
        t for cte in fp.ctes for t in (cte.get("source_tables") or [])
    })
    has_ctes = bool(fp.ctes)
    has_window = bool(re.search(r"\bover\s*\(", sql, re.IGNORECASE))
    n_joins = len(fp.joins)
    if has_ctes or has_window or n_joins >= 2:
        complexity = "complex"
    elif n_joins or len(fp.aggregations) >= 2 or len(fp.case_whens) >= 1:
        complexity = "medium"
    else:
        complexity = "simple"
    return {"tables": tables, "complexity": complexity, "parse_error": None}


# ─── Main extraction ─────────────────────────────────────────


def extract(
    file_path: Path,
    sheet_name: str | None,
    sql_col_name: str | None,
    prompt_col_name: str | None,
    difficulty_col_name: str | None,
    id_col_name: str | None,
    auto_detect_sql: bool,
    out_dir: Path,
    json_out: Path,
    use_gemini: bool,
    id_prefix: str = "Q",
    id_pad: int = 3,
) -> dict[str, Any]:
    wb = load_workbook_or_fail(file_path)
    sheets = list_sheets(wb)

    # Pick sheet
    if sheet_name:
        if sheet_name not in [s[0] for s in sheets]:
            wb.close()
            print(f"ERROR: sheet '{sheet_name}' not found. Available:", file=sys.stderr)
            for n, r, c in sheets:
                print(f"  - {n} ({r} rows × {c} cols)", file=sys.stderr)
            return {"status": "error"}
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[0]
        sheet_name = ws.title

    header = read_header(ws)
    if not header:
        wb.close()
        print(f"ERROR: sheet '{sheet_name}' has no header row", file=sys.stderr)
        return {"status": "error"}

    # Resolve SQL column
    sql_idx = col_index(header, sql_col_name)
    if sql_idx is None and auto_detect_sql:
        sql_idx = auto_pick_sql_column(ws, header)
        if sql_idx is not None:
            print(f"Auto-detected SQL column: '{header[sql_idx]}' (index {sql_idx})")
    if sql_idx is None:
        wb.close()
        print(
            "ERROR: SQL column not found. Pass --sql-col '<name>' (or --auto). "
            "Available headers: " + ", ".join(repr(h) for h in header),
            file=sys.stderr,
        )
        return {"status": "error"}

    prompt_idx = col_index(header, prompt_col_name)
    diff_idx = col_index(header, difficulty_col_name)
    id_idx = col_index(header, id_col_name)

    # Lazy Gemini client
    gemini_client = None
    if use_gemini:
        try:
            gemini_client = make_gemini_client()
        except Exception as e:
            print(
                f"ERROR: could not initialize Gemini client. Did you "
                f"`source agent_test/setup_vertex_env.sh`? {e}",
                file=sys.stderr,
            )
            wb.close()
            return {"status": "error"}

    out_dir.mkdir(parents=True, exist_ok=True)

    queries: list[dict[str, Any]] = []
    skipped_blank = 0
    skipped_not_sql = 0

    for row_n, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_n == 1:  # header
            continue
        raw_sql = _clean(row[sql_idx]) if sql_idx < len(row) else ""
        prompt = _clean(row[prompt_idx]) if prompt_idx is not None and prompt_idx < len(row) else ""
        difficulty = _clean(row[diff_idx]) if diff_idx is not None and diff_idx < len(row) else ""
        natural_id = _clean(row[id_idx]) if id_idx is not None and id_idx < len(row) else ""

        if not raw_sql and not prompt:
            skipped_blank += 1
            continue

        cleaned = basic_clean(raw_sql)
        cleaned_by_gemini = False
        if gemini_client and cleaned:
            cleaned, cleaned_by_gemini = gemini_clean(cleaned, gemini_client)
            if not cleaned:
                skipped_not_sql += 1
                continue

        meta = derive_metadata(cleaned) if cleaned else {
            "tables": [], "complexity": "unknown", "parse_error": "empty"
        }

        gen_id = natural_id or f"{id_prefix}{len(queries) + 1:0{id_pad}d}"
        record = {
            "id": gen_id,
            "prompt": prompt,
            "sql": cleaned,
            "difficulty": difficulty or None,
            "source_row": row_n,
            "source_sheet": sheet_name,
            "source_file": file_path.name,
            "tables": meta["tables"],
            "complexity": meta["complexity"],
            "parse_error": meta["parse_error"],
            "cleaned_by_gemini": cleaned_by_gemini,
        }
        queries.append(record)

        # Write the .sql file
        sql_path = out_dir / f"{gen_id}.sql"
        sql_path.write_text(cleaned + "\n", encoding="utf-8")

    wb.close()

    # Write the consolidated JSON
    json_out.parent.mkdir(parents=True, exist_ok=True)
    json_out.write_text(json.dumps(queries, indent=2), encoding="utf-8")

    return {
        "status": "ok",
        "extracted": len(queries),
        "skipped_blank": skipped_blank,
        "skipped_not_sql": skipped_not_sql,
        "sql_dir": str(out_dir),
        "json": str(json_out),
        "first_3": queries[:3],
    }


def main() -> int:
    parser = argparse.ArgumentParser(
        prog="excel_to_queries",
        description=(
            "Read SQL queries from an Excel column and produce data/gold_queries/"
            "Q*.sql + queries.json (the LUMI pipeline input)."
        ),
    )
    parser.add_argument("file", help="Path to .xlsx")
    parser.add_argument("--sheet", help="Sheet name (default: first)")
    parser.add_argument(
        "--sql-col",
        help="Header of the SQL column. If omitted, pass --auto or be prompted.",
    )
    parser.add_argument("--prompt-col", help="Optional NL question column")
    parser.add_argument("--difficulty-col", help="Optional difficulty column")
    parser.add_argument(
        "--id-col",
        help="Optional ID column. If absent, auto-generates Q001, Q002, ...",
    )
    parser.add_argument(
        "--auto",
        action="store_true",
        help="Auto-pick SQL column by heuristic (longest SQL-ish text)",
    )
    parser.add_argument(
        "--out-dir",
        default="data/gold_queries/",
        help="Where to write Q*.sql files",
    )
    parser.add_argument(
        "--json-out",
        default="data/gold_queries/queries.json",
        help="Where to write the consolidated JSON",
    )
    parser.add_argument(
        "--gemini-clean",
        action="store_true",
        help="Run each SQL through Gemini for normalization (one call per row).",
    )
    parser.add_argument("--id-prefix", default="Q")
    parser.add_argument("--id-pad", type=int, default=3)
    args = parser.parse_args()

    result = extract(
        file_path=Path(args.file),
        sheet_name=args.sheet,
        sql_col_name=args.sql_col,
        prompt_col_name=args.prompt_col,
        difficulty_col_name=args.difficulty_col,
        id_col_name=args.id_col,
        auto_detect_sql=args.auto,
        out_dir=Path(args.out_dir),
        json_out=Path(args.json_out),
        use_gemini=args.gemini_clean,
        id_prefix=args.id_prefix,
        id_pad=args.id_pad,
    )

    if result["status"] != "ok":
        return 1

    print()
    print(f"Extracted: {result['extracted']} queries")
    print(f"  skipped (blank rows):  {result['skipped_blank']}")
    print(f"  skipped (not SQL, Gemini-flagged): {result['skipped_not_sql']}")
    print(f"  .sql files:  {result['sql_dir']}/Q*.sql")
    print(f"  consolidated json: {result['json']}")
    print()
    print("First 3 records (preview):")
    for r in result["first_3"]:
        sql_preview = r["sql"][:100].replace("\n", " ")
        print(
            f"  [{r['id']}] tables={r['tables']} complexity={r['complexity']}\n"
            f"           prompt: {r.get('prompt', '')[:80]}\n"
            f"           sql:    {sql_preview}{'…' if len(r['sql']) > 100 else ''}"
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
