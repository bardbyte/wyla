"""Table catalog loader — scope + domain mapping from the table sheet.

Columns from the screenshot:
    table_name, IS IN DMP (Yes/No), company_domain, data_domain

The `IS IN DMP=Yes` flag is the scope signal — `Yes` rows are in
priority scope. The loader accepts xlsx via openpyxl if installed,
else CSV. Domain fields are normalized to strings (None when blank or
"Not Found"/"Not Available").
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable

from synapse.registry.schemas import TableCatalogEntry


_TABLE_NAME_KEYS = ("table_name", "table", "name", "table id")
_IS_IN_DMP_KEYS = ("is_in_dmp", "is in dmp", "dmp", "in_dmp", "in dmp")
_COMPANY_DOMAIN_KEYS = (
    "company_domain", "company domain", "domain", "department",
    "business_domain", "business domain",
)
_DATA_DOMAIN_KEYS = (
    "data_domain", "data domain", "data_category", "sub_domain",
)


def _pick(row: dict, candidates: tuple[str, ...]) -> str | None:
    lower = {(k or "").strip().lower(): v for k, v in row.items()}
    for c in candidates:
        v = lower.get(c)
        if v is not None and str(v).strip():
            return str(v).strip()
    return None


def _normalize_domain(s: str | None) -> str | None:
    """Treat 'Not Found' / 'Not Available' / blank as None."""
    if not s:
        return None
    cleaned = s.strip()
    if cleaned.lower() in {"not found", "not available", "n/a", "na", "-", ""}:
        return None
    return cleaned


def _parse_bool(s: str | None) -> bool:
    if not s:
        return False
    return s.strip().lower() in {"yes", "y", "true", "1", "in scope"}


def _iter_rows(path: Path) -> Iterable[dict]:
    """Stream rows from .xlsx or .csv. Header row becomes dict keys."""
    p = Path(path)
    suffix = p.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook  # type: ignore[import-not-found]
        except ImportError as e:
            raise RuntimeError(
                f"openpyxl required for xlsx input: {e}. "
                "Install with: pip install openpyxl",
            ) from e
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_raw = next(rows_iter)
        except StopIteration:
            return
        header = [
            str(h).strip() if h is not None else f"col_{i}"
            for i, h in enumerate(header_raw)
        ]
        for r in rows_iter:
            yield {header[i]: r[i] for i in range(min(len(header), len(r)))}
    else:
        with p.open(encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                yield dict(row)


def load_table_catalog(path: Path) -> list[TableCatalogEntry]:
    """Load every row of the table-catalog sheet."""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Table catalog not found: {p}")

    out: list[TableCatalogEntry] = []
    for row in _iter_rows(p):
        name = _pick(row, _TABLE_NAME_KEYS)
        if not name:
            continue
        out.append(TableCatalogEntry(
            table_name=name,
            is_in_dmp=_parse_bool(_pick(row, _IS_IN_DMP_KEYS)),
            company_domain=_normalize_domain(_pick(row, _COMPANY_DOMAIN_KEYS)),
            data_domain=_normalize_domain(_pick(row, _DATA_DOMAIN_KEYS)),
            raw_row=dict(row),
        ))
    return out


def in_scope(
    entries: list[TableCatalogEntry],
    *,
    require_dmp: bool = True,
    company_domains: set[str] | None = None,
) -> list[TableCatalogEntry]:
    """Filter to in-scope tables.

    `require_dmp=True` keeps only `IS IN DMP=Yes`.
    `company_domains={'Finance', ...}` restricts to those domains
    (case-insensitive).
    """
    out = entries
    if require_dmp:
        out = [e for e in out if e.is_in_dmp]
    if company_domains:
        wanted = {d.lower() for d in company_domains}
        out = [
            e for e in out
            if e.company_domain and e.company_domain.lower() in wanted
        ]
    return out


def index_by_domain(
    entries: list[TableCatalogEntry],
) -> dict[str, list[TableCatalogEntry]]:
    out: dict[str, list[TableCatalogEntry]] = {}
    for e in entries:
        key = (e.company_domain or "_unknown")
        out.setdefault(key, []).append(e)
    return out
