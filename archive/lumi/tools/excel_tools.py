"""Parse the gold-queries Excel into ParsedQuery objects (openpyxl + sqlglot).

No LLM. No regex-for-SQL. If sqlglot can't parse a row's SQL, we record the error
on the ParsedQuery and keep going — a partial corpus is better than a crash.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import sqlglot
from openpyxl import load_workbook
from sqlglot import exp

from lumi.schemas import Filter, JoinCondition, Measure, ParsedQuery

logger = logging.getLogger(__name__)

BQ_DIALECT = "bigquery"
_AGG_FUNCTIONS = {
    "SUM", "COUNT", "AVG", "MIN", "MAX", "STDDEV", "VAR", "MEDIAN",
    "COUNT_DISTINCT", "APPROX_DISTINCT", "ANY_VALUE", "ARRAY_AGG",
}


def parse_excel_to_json(
    file_path: str | Path,
    prompt_column: str = "user_prompt",
    sql_column: str = "expected_query",
    difficulty_column: str | None = "difficulty",
    sheet: str | None = None,
) -> dict[str, Any]:
    """Parse gold-queries Excel into structured ParsedQuery list.

    Args:
        file_path: Path to the .xlsx file.
        prompt_column: Column header for the natural-language prompt.
        sql_column: Column header for the expected SQL.
        difficulty_column: Column header for difficulty (optional).
        sheet: Sheet name; None picks the active/first sheet.

    Returns:
        dict with keys:
          status: "success" | "error"
          queries: list[ParsedQuery]  (empty on error)
          total_rows: int
          parse_errors: int
          error: str | None
    """
    path = Path(file_path)
    if not path.exists():
        return _err(f"Excel file not found: {path}")
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return _err(f"Unsupported extension {path.suffix}; expected .xlsx or .xlsm")

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        return _err(f"Failed to open workbook: {e}")

    ws = wb[sheet] if sheet else wb.active
    if ws is None:
        return _err("Workbook has no sheets.")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return _err("Empty worksheet.")

    header = [_to_str(c).strip() for c in header_row]
    col_idx = {name: i for i, name in enumerate(header) if name}

    for required in (prompt_column, sql_column):
        if required not in col_idx:
            return _err(
                f"Missing required column '{required}'. Found columns: {list(col_idx)}"
            )

    prompt_i = col_idx[prompt_column]
    sql_i = col_idx[sql_column]
    diff_i = col_idx.get(difficulty_column) if difficulty_column else None

    queries: list[ParsedQuery] = []
    parse_errors = 0
    total = 0

    for row_num, row in enumerate(rows_iter, start=2):
        prompt = _to_str(row[prompt_i] if prompt_i < len(row) else None).strip()
        sql = _to_str(row[sql_i] if sql_i < len(row) else None).strip()
        if not prompt and not sql:
            continue
        total += 1
        difficulty = (
            _to_str(row[diff_i]).strip() if diff_i is not None and diff_i < len(row) else None
        )

        query_id = f"q_{row_num:04d}"
        try:
            parsed = _parse_sql(query_id, prompt, sql, difficulty)
        except Exception as e:
            parse_errors += 1
            parsed = ParsedQuery(
                query_id=query_id,
                user_prompt=prompt,
                expected_sql=sql,
                difficulty=difficulty,
                parse_error=str(e),
            )
        queries.append(parsed)

    wb.close()

    logger.info(
        "Parsed %d queries from %s (%d parse errors)", len(queries), path, parse_errors
    )
    return {
        "status": "success",
        "queries": queries,
        "total_rows": total,
        "parse_errors": parse_errors,
        "error": None,
    }


def _parse_sql(query_id: str, prompt: str, sql: str, difficulty: str | None) -> ParsedQuery:
    tree = sqlglot.parse_one(sql, dialect=BQ_DIALECT)

    tables = _extract_tables(tree)
    primary_table = tables[0] if tables else None
    measures = _extract_measures(tree)
    dimensions = _extract_dimensions(tree)
    filters = _extract_filters(tree)
    joins = _extract_joins(tree)

    return ParsedQuery(
        query_id=query_id,
        user_prompt=prompt,
        expected_sql=sql,
        difficulty=difficulty,
        tables=tables,
        primary_table=primary_table,
        measures=measures,
        dimensions=dimensions,
        filters=filters,
        joins=joins,
    )


def _extract_tables(tree: exp.Expression) -> list[str]:
    seen: list[str] = []
    for t in tree.find_all(exp.Table):
        name = t.name
        if name and name not in seen:
            seen.append(name)
    return seen


def _extract_measures(tree: exp.Expression) -> list[Measure]:
    measures: list[Measure] = []
    for agg in tree.find_all(
        exp.Sum, exp.Count, exp.Avg, exp.Min, exp.Max, exp.Stddev, exp.Variance
    ):
        inner = agg.this
        distinct = False
        if isinstance(inner, exp.Distinct):
            distinct = True
            exprs = inner.expressions or [inner.this]
            inner = exprs[0] if exprs else None
        column = inner.name if isinstance(inner, exp.Column) else None
        measures.append(
            Measure(
                function=agg.__class__.__name__.upper(),
                column=column,
                distinct=distinct,
                expression=agg.sql(dialect=BQ_DIALECT),
            )
        )
    return measures


def _extract_dimensions(tree: exp.Expression) -> list[str]:
    dims: list[str] = []
    select = tree.find(exp.Select)
    if not select:
        return dims
    for e in select.expressions or []:
        unaliased = e.unalias() if isinstance(e, exp.Alias) else e
        if _is_aggregation(unaliased):
            continue
        if isinstance(unaliased, exp.Column) and unaliased.name not in dims:
            dims.append(unaliased.name)
    for g in tree.find_all(exp.Group):
        for col in g.expressions or []:
            if isinstance(col, exp.Column) and col.name not in dims:
                dims.append(col.name)
    return dims


def _extract_filters(tree: exp.Expression) -> list[Filter]:
    out: list[Filter] = []
    for where in tree.find_all(exp.Where):
        _flatten_predicates(where.this, out)
    return out


def _flatten_predicates(node: exp.Expression, out: list[Filter]) -> None:
    if isinstance(node, exp.And):
        _flatten_predicates(node.left, out)
        _flatten_predicates(node.right, out)
        return
    if isinstance(node, exp.Or):
        # capture both branches but flag via value that it's part of an OR
        _flatten_predicates(node.left, out)
        _flatten_predicates(node.right, out)
        return
    if isinstance(node, exp.EQ | exp.NEQ | exp.GT | exp.GTE | exp.LT | exp.LTE):
        col = _col_name(node.left)
        if col:
            out.append(Filter(column=col, operator=_op_symbol(node), value=node.right.sql(dialect=BQ_DIALECT)))
        return
    if isinstance(node, exp.In):
        col = _col_name(node.this)
        if col:
            vals = ", ".join(e.sql(dialect=BQ_DIALECT) for e in (node.expressions or []))
            out.append(Filter(column=col, operator="IN", value=f"({vals})"))
        return
    if isinstance(node, exp.Between):
        col = _col_name(node.this)
        if col:
            low = node.args.get("low")
            high = node.args.get("high")
            lo = low.sql(dialect=BQ_DIALECT) if low is not None else "?"
            hi = high.sql(dialect=BQ_DIALECT) if high is not None else "?"
            out.append(Filter(column=col, operator="BETWEEN", value=f"{lo} AND {hi}"))
        return
    if isinstance(node, exp.Like):
        col = _col_name(node.this)
        if col:
            out.append(Filter(column=col, operator="LIKE", value=node.expression.sql(dialect=BQ_DIALECT)))
        return
    if isinstance(node, exp.Is):
        col = _col_name(node.this)
        if col:
            out.append(Filter(column=col, operator="IS", value=node.expression.sql(dialect=BQ_DIALECT)))
        return


def _extract_joins(tree: exp.Expression) -> list[JoinCondition]:
    out: list[JoinCondition] = []
    for join in tree.find_all(exp.Join):
        right_tbl = join.this.name if isinstance(join.this, exp.Table) else None
        on = join.args.get("on")
        # sqlglot records LEFT/RIGHT/FULL on `side` and CROSS/INNER on `kind`.
        side = (join.side or "").lower()
        kind_raw = (join.kind or "").lower()
        kind = side or kind_raw or "inner"
        if right_tbl and isinstance(on, exp.EQ):
            left_col = on.left
            right_col = on.right
            lt = _col_table(left_col)
            lc = _col_name(left_col) or ""
            rt = _col_table(right_col) or right_tbl
            rc = _col_name(right_col) or ""
            if lt and rt and lc and rc:
                out.append(
                    JoinCondition(
                        left_table=lt,
                        left_column=lc,
                        right_table=rt,
                        right_column=rc,
                        join_type=kind,
                    )
                )
    return out


def _is_aggregation(node: exp.Expression) -> bool:
    return any(
        isinstance(node, cls)
        for cls in (exp.Sum, exp.Count, exp.Avg, exp.Min, exp.Max, exp.Stddev, exp.Variance)
    ) or (
        isinstance(node, exp.Anonymous)
        and (node.name or "").upper() in _AGG_FUNCTIONS
    )


def _col_name(node: exp.Expression | None) -> str | None:
    if isinstance(node, exp.Column):
        return node.name
    return None


def _col_table(node: exp.Expression | None) -> str | None:
    if isinstance(node, exp.Column) and node.table:
        return node.table
    return None


def _op_symbol(node: exp.Expression) -> str:
    return {
        exp.EQ: "=",
        exp.NEQ: "!=",
        exp.GT: ">",
        exp.GTE: ">=",
        exp.LT: "<",
        exp.LTE: "<=",
    }.get(type(node), "?")


def _to_str(v: object) -> str:
    return "" if v is None else str(v)


def _err(msg: str) -> dict[str, Any]:
    logger.error(msg)
    return {
        "status": "error",
        "queries": [],
        "total_rows": 0,
        "parse_errors": 0,
        "error": msg,
    }
